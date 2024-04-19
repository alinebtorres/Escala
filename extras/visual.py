import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from tempfile import NamedTemporaryFile
from shutil import copyfileobj

def identificar_escala(excel_file):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    turnos = {
        "M15": "Manhã",
        "M17": "Manhã",
        "T15": "Tarde",
        "D6": "Dia",
        "N8": "Noite",
        "AB": "Abono",
        "CH-12": "Folga"
    }
    escalas_por_dia = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        funcionario = row[0]
        for i, turno in enumerate(row[1:], start=1):
            if turno in turnos:
                dia = i
                if dia not in escalas_por_dia:
                    escalas_por_dia[dia] = []
                escalas_por_dia[dia].append((funcionario, turnos[turno]))
    return escalas_por_dia

def organizar_turnos(escalas_por_dia):
    funcionarios_organizados = []
    prioridades = {
        "Dia": 1,
        "Manhã": 2,
        "Tarde": 3,
        "Noite": 4,
        "CH-12": 5,
        "Abono": 6
    }
    for dia in range(1, 32):
        if dia in escalas_por_dia:
            turnos_do_dia = sorted(escalas_por_dia[dia], key=lambda x: prioridades.get(x[1], float('inf')))
            for funcionario, turno in turnos_do_dia:
                funcionarios_organizados.append((dia, funcionario, turno))
    return funcionarios_organizados

def salvar_em_excel(funcionarios_organizados, output_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=1, value="Dia")
    sheet.cell(row=1, column=2, value="Funcionário")
    sheet.cell(row=1, column=3, value="Turno")
    current_number = None
    current_row = 1
    for dia, funcionario, turno in funcionarios_organizados:
        if dia != current_number:
            current_number = dia
            current_row += 1
            sheet.insert_rows(current_row)
        current_row += 1
        sheet.cell(row=current_row, column=1, value=dia)
        sheet.cell(row=current_row, column=2, value=funcionario)
        sheet.cell(row=current_row, column=3, value=turno)
    workbook.save(output_file)

def selecionar_arquivo():
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if filename:
        entry.delete(0, tk.END)
        entry.insert(tk.END, filename)

def processar_e_salvar():
    excel_file = entry.get()
    if not excel_file:
        messagebox.showerror("Erro", "Por favor, selecione um arquivo Excel.")
        return

    try:
        escalas_por_dia = identificar_escala(excel_file)
        funcionarios_organizados = organizar_turnos(escalas_por_dia)

        with NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            salvar_em_excel(funcionarios_organizados, temp_file.name)
        
        output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if output_file:
            copyfileobj(open(temp_file.name, 'rb'), open(output_file, 'wb'))

        messagebox.showinfo("Concluído", "As escalas foram organizadas e salvas com sucesso!")
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao processar o arquivo: {e}")


root = ctk.CTk()
root.title("Organizador de Escalas")
root.geometry('500x350')
label = ctk.CTkLabel(root, text='Escala de Atribuição', font=('Arial', 20, 'bold'))
label.pack(pady=20, padx=5)


root.configure(background='green')

select_button = ctk.CTkButton(root, text="Selecionar Arquivo", command=selecionar_arquivo)
select_button.pack(pady=15)

entry = ctk.CTkEntry(root, width=140)
entry.pack(pady=10, padx=15)



process_button = ctk.CTkButton(root, text="Processar e Salvar", command=processar_e_salvar)
process_button.pack(pady=5)



root.mainloop()