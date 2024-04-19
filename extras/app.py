import openpyxl

def identificar_escala(excel_file):
    # Abra o arquivo do Excel
    workbook = openpyxl.load_workbook(excel_file)
    
    # Selecione a primeira planilha
    sheet = workbook.active
    
    # Dicionário para mapear os códigos de turno para seus respectivos significados
    turnos = {
        "M15": "Manhã",
        "M17": "Manhã",
        "T15": "Tarde",
        "D6": "Dia",
        "N8": "Noite",
        "AB": "Abono",
        "CH-12": "Folga"
    }
    
    # Lista para armazenar as informações de cada dia
    escalas_por_dia = {}
    
    # Iterar sobre as células da planilha (ignorando a primeira linha que contém os nomes dos funcionários)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        funcionario = row[0]  # Nome do funcionário na primeira coluna
        
        # Iterar sobre os dias do mês
        for i, turno in enumerate(row[1:], start=1):
            # Verificar se o valor da célula é um código de turno
            if turno in turnos:
                # Obter o dia correspondente
                dia = i
                if dia not in escalas_por_dia:
                    escalas_por_dia[dia] = []
                escalas_por_dia[dia].append((funcionario, turnos[turno]))
    
    # Retornar as informações de cada dia
    return escalas_por_dia

def organizar_turnos(escalas_por_dia):
    # Lista para armazenar os funcionários organizados
    funcionarios_organizados = []

    # Prioridades dos turnos
    prioridades = {
        "Dia": 1,
        "Manhã": 2,
        "Tarde": 3,
        "Noite": 4,
        "CH-12": 5,
        "Abono": 6
    }

    # Organizar os funcionários por dia e prioridade de turno
    for dia in range(1, 32):  # Iterar sobre os dias do mês
        if dia in escalas_por_dia:
            # Ordenar os turnos por prioridade
            turnos_do_dia = sorted(escalas_por_dia[dia], key=lambda x: prioridades.get(x[1], float('inf')))
            for funcionario, turno in turnos_do_dia:
                funcionarios_organizados.append((dia, funcionario, turno))

    return funcionarios_organizados

def salvar_em_excel(funcionarios_organizados, output_file):
    # Criar um novo arquivo Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # Escrever os dados na nova planilha
    sheet.cell(row=1, column=1, value="Dia")
    sheet.cell(row=1, column=2, value="Funcionário")
    sheet.cell(row=1, column=3, value="Turno")

    current_number = None
    current_row = 1

    for dia, funcionario, turno in funcionarios_organizados:
        if dia != current_number:
            current_number = dia
            current_row += 1
            sheet.insert_rows(current_row)

        current_row += 1
        sheet.cell(row=current_row, column=1, value=dia)
        sheet.cell(row=current_row, column=2, value=funcionario)
        sheet.cell(row=current_row, column=3, value=turno)

    # Salvar o novo arquivo Excel
    workbook.save(output_file)

# Chamar a função para identificar a escala
escalas_por_dia = identificar_escala("esc.xlsx")

# Chamar a função para organizar os turnos
funcionarios_organizados = organizar_turnos(escalas_por_dia)

# Salvar as informações em uma nova planilha Excel
salvar_em_excel(funcionarios_organizados, "escalas_organizadas.xlsx")